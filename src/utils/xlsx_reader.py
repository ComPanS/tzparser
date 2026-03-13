"""
Чтение ИНН из xlsx файла.

Поддерживает настраиваемое имя листа и колонки через конфиг.
"""
import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from src.config import XLSX_INN_COLUMN, XLSX_SHEET_NAME

logger = logging.getLogger(__name__)

# ИНН: 10 цифр (юрлицо) или 12 цифр (физлицо/ИП)
INN_PATTERN = re.compile(r"^\d{10}$|^\d{12}$")


def _normalize_inn(value: str) -> str:
    """Убирает пробелы из ИНН."""
    return re.sub(r"\s+", "", str(value).strip())


def _is_valid_inn(inn: str) -> bool:
    """Проверяет формат ИНН (10 или 12 цифр)."""
    return bool(INN_PATTERN.match(inn)) if inn else False


def read_inns_from_xlsx(
    path,
    sheet_name=None,
    column=None,
):
    """
    Читает список ИНН из xlsx файла.

    Args:
        path: Путь к xlsx файлу
        sheet_name: Имя листа (по умолчанию из конфига)
        column: Имя колонки с ИНН (по умолчанию из конфига)

    Returns:
        Список валидных ИНН (без дублей)
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    sheet_name = sheet_name or XLSX_SHEET_NAME
    column = column or XLSX_INN_COLUMN

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист '{sheet_name}' не найден. "
                f"Доступные: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]
    finally:
        wb.close()

    inns = []
    seen = set()
    invalid_count = 0

    # Ищем колонку по заголовку в первой строке
    headers = [cell.value for cell in ws[1]]
    col_idx = None
    for i, h in enumerate(headers):
        if h and str(h).strip() == column:
            col_idx = i
            break

    if col_idx is None:
        raise ValueError(
            f"Колонка '{column}' не найдена. "
            f"Заголовки: {headers}"
        )

    for row in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1):
        cell = row[0]
        if cell.value is None:
            continue
        raw = str(cell.value).strip()
        inn = _normalize_inn(raw)
        if not inn:
            continue
        if not _is_valid_inn(inn):
            invalid_count += 1
            logger.warning("Некорректный ИНН (пропущен): %s", raw)
            continue
        if inn not in seen:
            seen.add(inn)
            inns.append(inn)

    if invalid_count:
        logger.info("Пропущено некорректных ИНН: %d", invalid_count)
    logger.info("Загружено ИНН: %d", len(inns))
    return inns
