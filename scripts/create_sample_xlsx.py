"""Создаёт пример xlsx с ИНН для тестирования."""
from pathlib import Path

from openpyxl import Workbook

# ИНН из ТЗ (231138771115) + несколько тестовых
SAMPLE_INNS = [
    "231138771115",
    "7707083893",
    "7736207543",
]

ROOT = Path(__file__).parent.parent
OUTPUT = ROOT / "data" / "inns.xlsx"


def main():
    (ROOT / "data").mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ИНН"
    for i, inn in enumerate(SAMPLE_INNS, start=2):
        ws[f"A{i}"] = inn
    wb.save(OUTPUT)
    print(f"Создан {OUTPUT}")


if __name__ == "__main__":
    main()
